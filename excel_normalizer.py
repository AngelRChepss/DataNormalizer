from __future__ import annotations
from typing import List, Dict, Tuple, Iterable, Any, Callable
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
from text_normalizer import Normalizer
from norm_utils import check_rut_normalize, validate_email_strict

class SheetNormalizer:
    FILL_NORMALIZED = PatternFill(fill_type="solid", fgColor="FFCCFFFF")
    FILL_INVALID = PatternFill(fill_type="solid", fgColor="FFFF4444")
    FILL_UNMAPPED = PatternFill(fill_type="solid", fgColor="FFFFBB99")
    FILL_NOTFOUND = PatternFill(fill_type="solid", fgColor="FF6666FF")
    FILL_TOOMANY = PatternFill(fill_type="solid", fgColor="FFFF8888")
    FILL_DUPLICATE = PatternFill(fill_type="solid", fgColor="FFAAAA55")
    FONT_BASE = Font(bold=False)

    def __init__(self, worksheet: Worksheet, wb_normalizer: BookNormalizer):
        self.ws = worksheet
        self._max_row: int = None
        self._header_map = None
        self._max_column: int = None
        self.wb_normalizer = wb_normalizer

    def recalculate_header_map(self):
        self._header_map = {
            cell.value: cell.column_letter
            for idx, cell in enumerate(self.ws[1], start=1)
            if cell.value
        }

    @property
    def header_map(self) -> Dict[str, str]:
        if self._header_map is None:
            self.recalculate_header_map()
        return self._header_map

    def recalculate_max_row(self):
        max_row = self.ws.max_row
        while max_row > 0 and all(cell.value is None for cell in self.ws[max_row]):
            max_row -= 1
        self._max_row = max_row

    @property
    def max_row(self) -> int:
        if self._max_row is None:
            self.recalculate_max_row()
        return self._max_row

    def recalculate_max_column(self):
        self._max_column = len([cell for cell in self.ws[1] if cell.value])

    @property
    def max_column(self) -> int:
        if self._max_column is None:
            self.recalculate_max_column()
        return self._max_column

    def header_map_cols(self, *cols) -> List[str]:
        if cols:
            return [self.header_map[col] for col in cols]
        else:
            return list(self.header_map.values())

    def col_to_letter(self, col) -> str:
        if isinstance(col, int):
            col_letter = get_column_letter(col)
        else:
            col_letter = self.header_map.get(str(col), str(col))
        return col_letter

    def __getitem__(self, key):
        col, row = key
        return self.ws[f"{self.col_to_letter(col)}{row}"].value

    def __setitem__(self, key, value):
        col, row = key
        col_letter = self.col_to_letter(col)
        self.ws[f"{col_letter}{row}"].value = value
        if row == 1:
            self.recalculate_max_column()
            self.recalculate_header_map()
        if row > self.max_row:
            self.recalculate_max_row()

    def get_row(self, row: int, *cols : str) -> Tuple:
        if len(cols) == 0 or not cols:
            cols = range(1, self.max_column + 1)
        return tuple(self[col, row] for col in cols)

    def paint(self, col: str | int, row: int, pattern: PatternFill) -> None:
        self.ws[f"{self.col_to_letter(col)}{row}"].fill = pattern

    def comment_cell(self, col: str | int, row: int, comment: str):
        col = self.col_to_letter(col)
        self.ws[f"{col}{row}"].comment = Comment(comment, "normalizer")

    @staticmethod
    def change_cell(cell: Cell, value, pattern: PatternFill | None = None, font: Font | None = None):
        cell.value = value
        if pattern:
            cell.fill = pattern
        if font:
            cell.font = font

    def normalize_columns(self, columns: List[str], normalizer: Normalizer, start_row: int = 2) -> None:
        """
        Normaliza una lista de columnas de TEXTO en un Worksheet
        """
        for column in columns:
            cells = self.ws[self.header_map[column]]
            for cell in cells:
                if cell.row < start_row:
                    continue
                # Normalizer normaliza None a "", porque espera strings.
                # Pero nosotros preferimos quedarnos con None. Más aún, textos vacíos
                # también deben ser None.
                if not (cell.value is None):
                    result = normalizer.normalize(str(cell.value))
                    if cell.value != result:
                        # Cambia texto vacío a None
                        SheetNormalizer.change_cell(cell, result or None, pattern=SheetNormalizer.FILL_NORMALIZED)

    def find_uniques(self, column: str, exclude_empty: bool = True, sort: bool = False, start_row : int = 2) -> List:
        """
        Encuentra todos los valores unicos en una columna de valores y retorna una lista con ellos.
        La columna puede tener cualquier tipo de datos.
        """
        cells = self.ws[self.header_map[column]]

        values = {
            cell.value for cell in cells
            if cell.row >= start_row and not (exclude_empty and cell.value in (None, ""))
        }
        values = sorted(values) if sort else list(values)
        return values

    def find_multicolumn_uniques(self, columns: Iterable[str], sort: bool = False, start_row: int = 2) -> List[Tuple]:
        """
        Encuentra todos los valores unicos en una columna de valores y retorna una lista con ellos.
        La columna puede tener cualquier tipo de datos.
        """
        values = set()
        cols = self.header_map_cols(*columns)
        for row in range(start_row, self.max_row + 1):
            values.add(tuple(self.ws[f"{col}{row}"].value or "" for col in cols))
        values = sorted(values) if sort else list(values)
        return values

    def highlight_invalid_ruts(self, column: str) -> int:
        cells = self.ws[self.header_map[column]]
        invalid_count = 0
        for cell in cells:
            if cell.value:
                valid, norm = check_rut_normalize(str(cell.value))
                if valid:
                    continue
            cell.fill = SheetNormalizer.FILL_INVALID
            invalid_count += 1
        return invalid_count

    def normalize_ruts(self, column: str, norm_mode="standard", validation_mode="strict", start_row: int = 2) -> int:
        cells = self.ws[self.header_map[column]]
        invalid_count = 0
        for cell in cells:
            if cell.row < start_row:
                continue
            if cell.value:
                valid, norm, msg = check_rut_normalize(str(cell.value), norm_mode=norm_mode, validation_mode=validation_mode)
                if valid:
                    if norm != cell.value:
                        SheetNormalizer.change_cell(cell, norm, pattern=SheetNormalizer.FILL_NORMALIZED)
                    continue
                else:
                    cell.fill = SheetNormalizer.FILL_INVALID
                    cell.comment = Comment(f"Rut invalido: {msg}", "normalizer")
                    invalid_count += 1
            else:
                cell.fill = SheetNormalizer.FILL_INVALID
                cell.comment = Comment(f"Rut invalido: Campo nulo", "normalizer")
                invalid_count += 1
        return invalid_count

    def write_values(self, values: Dict) -> None:
        """
        Agrega todos los datos entregados como columnas.
        """
        col_idx = self.max_column + 1
        for column_name, column_values in values.items():
            col_letter = get_column_letter(col_idx)
            self.ws[col_letter + "1"].value = column_name
            row = 2
            for value in column_values:
                self.ws[col_letter + str(row)].value = value
                row += 1
            col_idx += 1
        self.recalculate_max_row()
        self.recalculate_header_map()
        self.recalculate_max_column()

    def map_cols_unsafe(self, mapping_function, *cols) -> None:
        cols = self.header_map_cols(*cols)
        for col in cols:
            for row in range(2, self.max_row + 1):
                cell = self.ws[f"{col}{row}"]
                cell.value = mapping_function(cell.value)

    def map_cols_safe(self, mapping_function, *cols) -> None:
        cols = self.header_map_cols(*cols)
        for col in cols:
            for row in range(2, self.max_row + 1):
                cell = self.ws[f"{col}{row}"]
                try:
                    cell.value = mapping_function(cell.value)
                except Exception as e:
                    cell.fill = SheetNormalizer.FILL_INVALID
                    cell.comment = Comment(str(e), "normalizer")

    def multimap_cols_unsafe(self, mapping_function, *cols) -> None:
        header_map = self.header_map_cols(*cols)
        for row in range(2, self.max_row + 1):
            map_data = tuple(
                self[col,row] for col in header_map
            )
            new_data = mapping_function(*map_data)
            for i, col in enumerate(cols):
                self[col,row] = new_data[i]

    def get_columns(self, *cols : str) -> Dict[str, List]:
        data = {}
        for col in cols:
            data[col] = []
            for row in range(2, self.max_row + 1):
                data[col].append(self[col, row])
        return data

    def map_with_dict(self, mapper: Dict, column: str, tgt_column: str):
        maps = mapper.keys()
        values = mapper.values()
        for r in range(2, self.max_row + 1):
            value = self[column, r]
            self[tgt_column, r] = mapper.get(value, value)
            if value not in maps and value not in values:
                self.paint(tgt_column, r, self.FILL_UNMAPPED)

    def look_up(self, compare_value, lookup_cols: Iterable = None,  comparer: Callable[[Tuple, Any], bool] = None) -> List[Tuple]:
        comparer = comparer or (lambda x, y : x == y)
        results = []
        lookup_cols = lookup_cols or self.header_map_cols()
        for row in range(2, self.max_row + 1):
            row_values = (row, ) + tuple(self[col, row] for col in lookup_cols)
            if comparer(compare_value, row_values):
                results.append(row_values)
        return results

    def overwrite_rows(self, *rows):
        for idx, row in enumerate(rows):
            for col in range(1, self.max_column + 1):
                # +1 para compensar porque excel parte de 1, y +1 para saltar header
                self[col, idx+2] = row[col-1]

    def sort_columns(self, *cols):
        # Iterar para crear una lista de tuplas de filas
        data = []
        for row in range(2, self.max_row + 1):
            data.append(self.get_row(row))
        index_cols = [column_index_from_string(self.col_to_letter(col)) for col in cols]
        # Ordenar por cada columna por separado
        for col in reversed(index_cols):
            data.sort(key = lambda x: (x[col-1] is not None, str(x[col-1]) if x[col-1] is not None else ""), reverse = False)
        self.overwrite_rows(*data)

    def create_column(self, name: str) -> None:
        self[self.max_column + 1, 1] = name
        self.recalculate_max_column()
        self.recalculate_header_map()

    def highlight_duplicates(self, column):
        """Destaca todos los valores duplicados en una columna especificada"""
        data = {}
        for row in range(2, self.max_row + 1):
            value = self[column, row]
            if value in data.keys():
                data[value].append(row)
                for dup in data[value]:
                    self.paint(column, dup, self.FILL_DUPLICATE)
                    self.comment_cell(column, dup, f"Valor duplicado en {data[value]}")
            else:
                data[value] = [row]

    def normalize_emails(self, column: str, normalizer: Normalizer = None):
        """Normaliza todos los emails de una columna"""
        for row in range(2, self.max_row + 1):
            value = self[column, row]
            if value and normalizer:
                value = normalizer.normalize(value)
            if value:
                self[column, row] = value
                valid, msg = validate_email_strict(value)
                if not valid:
                    self.paint(column, row, self.FILL_INVALID)
                    self.comment_cell(column, row, msg)
            else:
                self.paint(column, row, self.FILL_INVALID)
                self.comment_cell(column, row, "Campo vacío")

    def split_column(self, source_col: str, new_cols: list[str], delimiter: str, start_row: int = 2):
        """
        Split a source column into multiple new columns using a delimiter.

        Example:
            split_column("Full Name", ["First", "Middle", "Last", "Suffix"], " ")
        """
        # Ensure new columns exist
        for name in new_cols:
            self.create_column(name)

        for row in range(start_row, self.max_row + 1):
            value = self[source_col, row]
            if value is None:
                continue
            parts = str(value).split(delimiter, len(new_cols) - 1)
            for i, col_name in enumerate(new_cols):
                self[col_name, row] = parts[i] if i < len(parts) else None

    def copy_column(self, source_col: str, new_col: str):
        # Create new column with given name
        self.create_column(new_col)
        tgt_letter = self.header_map[new_col]
        src_letter = self.header_map[source_col]

        for row in range(2, self.max_row + 1):
            value = self.ws[f"{src_letter}{row}"].value
            self.ws[f"{tgt_letter}{row}"].value = value

class BookNormalizer:
    def __init__(self, file_name: str):
        self.wb: Workbook = load_workbook(file_name)
        self.ws_norms = {sheet: SheetNormalizer(self.wb[sheet], self) for sheet in self.wb.sheetnames}
        self.current_norm = self.ws_norms[self.wb.sheetnames[0]]
        self.mappings : Dict[str, Dict] = {}
        self.file_name = file_name

    def keep_sheets(self, sheets: Iterable | None = None) -> None:
        sheets = sheets or self.wb.sheetnames
        wb_sheets = self.wb.sheetnames.copy()
        for sheet in wb_sheets:
            if sheet not in sheets:
                del self.wb[sheet]

    def save(self, file_name: str) -> None:
        self.wb.save(file_name)

    def create_sheet(self, sheet_name: str) -> None:
        ws = self.wb.create_sheet(sheet_name)
        self.ws_norms[sheet_name] = SheetNormalizer(ws, self)

    def join_columns(self, target_worksheet: str, columns: List[str], target_name: str, join_character: str = " ") -> None:
        """
        Combina varias columnas en una sola que escribe en otra hoja.
        """
        cols = [self.header_map[col] for col in columns]
        max_row = self.max_row
        tgt_wsn = self.ws_norms[target_worksheet]

        values = []
        for row in range(2, max_row+1):
            data = []
            for col in cols:
                data.append(self.ws[f"{col}{row}"].value)
            values.append(join_character.join(str(x) for x in data if x is not None and x != "").strip())

        tgt_wsn.write_values({target_name: values})

    def activate_sheet(self, sheet_name: str) -> None:
        self.current_norm = self.ws_norms[sheet_name]

    def unify_into_sheet(self, column: str, new_name: str, target_sheet: str, exclude_empty: bool = True, sort: bool = False, start_row : int = 2):
        unified = self.sheet.find_uniques(column, exclude_empty, sort, start_row)
        data = {new_name: unified}
        self.ws_norms[target_sheet].write_values(data)

    def multi_unify_into_sheet(self, columns: Iterable[str], new_names: Iterable[str], target_sheet: str, sort: bool = False, start_row : int = 2):
        unified = self.sheet.find_multicolumn_uniques(columns, sort, start_row)
        data = {
            name: []
            for name in new_names
        }

        for uni in unified:
            for i, unique in enumerate(uni):
                data[new_names[i]].append(unique)

        self.ws_norms[target_sheet].write_values(data)

    def copy_cols_into_sheet(self, target_sheet: str, *cols: str) -> None:
        data = self.current_norm.get_columns(*cols)
        self.ws_norms[target_sheet].write_values(data)

    def save_sheets_to_file(self, file_name: str, *sheet_names: str,):
        self.wb.save(file_name)
        bn = BookNormalizer(file_name)
        bn.keep_sheets(sheet_names)
        bn.save(file_name)
        bn.wb.close()


    @property
    def sheet(self) -> SheetNormalizer:
        """Acceso explícito al SheetNormalizer actual (para autocompletado)."""
        return self.current_norm

    def load_mapping(self, sheet : str, key_col: str, value_col: str, mapping_name: str = None, file: str = ""):
        close = False
        if (file or self.file_name) != self.file_name:
            wb = load_workbook(file)
            ws = wb[sheet]
            normal = SheetNormalizer(ws, self)
            close = True
        else:
            wb = None
            normal = self.ws_norms[sheet]
        data = normal.get_columns(key_col, value_col)
        if close:
            wb.close()

        mapped = {}
        for row in range(len(data[key_col])):
            mapped[data[key_col][row]] = data[value_col][row]
        self.mappings[mapping_name or sheet] = mapped

    def apply_mapping(self, mapping_name: str, column, tgt_column) -> None:
        mapper = self.mappings[mapping_name]
        self.current_norm.map_with_dict(mapper, column, tgt_column)

    def lookup_map(self,
                   mapper: Callable[[Tuple, Tuple], Any],
                   mapping_cols: List[str],
                   comparer : Callable[[Tuple, Tuple], bool],
                   lookup_cols : List[str],
                   look_up_sheet : str
                   ):
        """TODO: Documentar porque es muy compleja de usar!"""
        lookup_norm = self.ws_norms[look_up_sheet]
        # Iterar sobre filas en columnas
        for row in range(2, self.current_norm.max_row + 1):
            row_data = (row,) + (self.current_norm.get_row(row, *mapping_cols))
            search_result = lookup_norm.look_up(row_data, lookup_cols, comparer)
            if not search_result:
                for col in mapping_cols:
                    self.current_norm.paint(col, row, self.current_norm.FILL_NOTFOUND)
                continue
            elif len(search_result) > 1:
                for col in mapping_cols:
                    self.current_norm.paint(col, row, self.current_norm.FILL_TOOMANY)
                    self.current_norm.comment_cell(col, row, f"Found multiple: {search_result}")
                # Mapeamos el primer hallazgo de todos modos
            search_result = search_result[0]
            result = mapper(row_data, search_result)
            for i, value in enumerate(result):
                self.current_norm[mapping_cols[i], row] = value

    def merge_columns_into_sheet(
            self,
            column_groups: list[tuple[str, ...]],
            target_sheet: str,
            new_names: tuple[str, ...],
            *,
            drop_empty_rows: bool = True,
            dedupe: bool = False,
    ):
        """
        Merge rows from multiple source column groups into a new sheet with unified column names.

        Args:
            column_groups: e.g. [
                ("ColA1", "ColB1"),
                ("ColA2", "ColB2"),
                # ...more groups, each same length as new_names
            ]
            target_sheet: sheet name to write to (created if missing)
            new_names: output column headers, e.g. ("ColA", "ColB")
            drop_empty_rows: skip rows where all values are empty/None
            dedupe: keep only unique row-tuples across the merged result
        """
        k = len(new_names)
        if target_sheet not in self.ws_norms:
            self.create_sheet(target_sheet)  # creates SheetNormalizer for it

        # Collect merged data per output column
        merged = {name: [] for name in new_names}

        for group in column_groups:
            if len(group) != k:
                raise ValueError(f"Column group length {len(group)} != {k} (len(new_names))")

            # Read all source columns (returns lists without headers)
            cols_dict = self.sheet.get_columns(*group)  # uses column names as keys
            # Reconstruct row-wise tuples for this group
            rows_iter = zip(*(cols_dict[col] for col in group))

            for row_vals in rows_iter:
                if drop_empty_rows and all(v in (None, "") for v in row_vals):
                    continue
                for i, out_name in enumerate(new_names):
                    merged[out_name].append(row_vals[i])

        if dedupe:
            seen = set()
            deduped = {name: [] for name in new_names}
            for row in zip(*(merged[name] for name in new_names)):
                if row in seen:
                    continue
                seen.add(row)
                for i, name in enumerate(new_names):
                    deduped[name].append(row[i])
            merged = deduped

        # Write merged columns to the target sheet
        self.ws_norms[target_sheet].write_values(merged)

    def close_book(self):
        self.wb.close()

    def __getattr__(self, name):
        """
        Delegar metodos y atributos al SheetNormalizer actual.
        Mejora mantenibilidad de las clases.
        """
        if self.current_norm and hasattr(self.current_norm, name):
            return getattr(self.current_norm, name)
        raise AttributeError(f"'BookNormalizer' object has no attribute '{name}'")
